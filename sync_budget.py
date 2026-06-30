#!/usr/bin/env python3
"""
sync_budget.py — DCR Budget Sync
=================================
Reads real CSV exports from:
  ./Amex/    → Amex BA Black Card (activity*.csv)
  ./Monzo/   → Monzo current account + Flex card (MonzoDataExport*.csv)

Updates DCR_Budget_Tracker.xlsx → Transactions sheet.
Sends budget alert emails at 80% threshold via Gmail.

USAGE:
  python3 sync_budget.py           → sync only
  python3 sync_budget.py --brief   → sync + send monthly CFO brief
  python3 sync_budget.py --check   → alerts check only

FIRST-TIME SETUP:
  1. Open this file and fill in GMAIL_USER and GMAIL_APP_PASSWORD below
     (Get an App Password at: myaccount.google.com/apppasswords)
  2. Run: python3 sync_budget.py
  3. Open DCR_Budget_Tracker.xlsx — all transactions will be there
"""

import os, sys, csv, re, hashlib, smtplib, argparse, json
from datetime import date, datetime
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── USER CONFIG ──────────────────────────────────────────────────────────────

BASE_DIR      = Path(__file__).parent
XLSX_PATH     = BASE_DIR / "DCR_Budget_Tracker.xlsx"
MONZO_DIR     = BASE_DIR / "Monzo"

ALERT_EMAIL      = "daniel.cruz.rosso@gmail.com"
ALERT_THRESHOLD  = 0.80   # 80%

# Gmail SMTP — fill these in once
GMAIL_USER         = "daniel.cruz.rosso@gmail.com"
GMAIL_APP_PASSWORD = ""   # Get at: myaccount.google.com/apppasswords

# ─── CATEGORIES & MERCHANT MAP ────────────────────────────────────────────────

CATEGORIES = [
    "Dining", "Groceries", "Health", "Transport",
    "Shopping", "Entertainment", "Subscriptions",
    "Personal care", "Travel", "Misc", "Rent",
]

MERCHANT_MAP = {
    # Rent
    "savills": "Rent",
    # Dining
    "deliveroo": "Dining", "uber eats": "Dining", "just eat": "Dining",
    "dishoom": "Dining", "pizza": "Dining", "restaurant": "Dining",
    "cafe": "Dining", "coffee": "Dining", "starbucks": "Dining",
    "costa": "Dining", "pret": "Dining", "mcdonalds": "Dining",
    "nando": "Dining", "wagamama": "Dining", "salad": "Dining",
    "dumpling": "Dining", "patisserie": "Dining", "guinea": "Dining",
    "climpson": "Dining", "digby": "Dining", "teresa": "Dining",
    "zettle": "Dining",   # Zettle card reader = typically cafe/restaurant
    "bullgogi": "Dining", "black bear": "Dining", "kati roll": "Dining",
    "brilliant corners": "Dining", "albion": "Dining", "gunmaker": "Dining",
    "cock and bottle": "Dining", "gail": "Dining", "oseyo": "Dining",
    "jcb kanazawa": "Dining", "dojo*": "Dining", "crown bow": "Dining",
    "tst-": "Dining",     # Toast POS prefix = usually bar/restaurant
    "five guys": "Dining", "barrafina": "Dining", "wetherspoon": "Dining",
    "marlborough": "Dining", "pelican": "Dining", "pub": "Dining",
    "arms": "Dining", "bar": "Dining", "inn": "Dining", "tavern": "Dining",
    "casa": "Dining",
    # Entertainment / Leisure
    "soho house": "Entertainment", "pitchgolf": "Entertainment",
    "cinema": "Entertainment", "odeon": "Entertainment", "vue": "Entertainment",
    "ticketmaster": "Entertainment", "eventbrite": "Entertainment",
    "theatre": "Entertainment", "bowling": "Entertainment",
    "gocardless": "Entertainment", "golf": "Entertainment",
    # Groceries
    "tesco": "Groceries", "sainsbury": "Groceries", "waitrose": "Groceries",
    "m&s": "Groceries", "ms city": "Groceries", "ocado": "Groceries",
    "lidl": "Groceries", "aldi": "Groceries", "whole foods": "Groceries",
    "co-op": "Groceries", "morrisons": "Groceries",
    # Health
    "gym": "Health", "fitness": "Health", "pharmacy": "Health",
    "boots": "Health", "dentist": "Health", "dental": "Health",
    "doctor": "Health", "physio": "Health", "nuffield": "Health",
    "barry": "Health", "equinox": "Health", "well being": "Health",
    "ryft": "Health", "index medical": "Health", "medical": "Health",
    "jimmy fairly": "Health",   # optician
    "stretch life": "Health",   # yoga/stretch studio
    # Shopping
    "best buy": "Shopping",
    # Transport (commuting, public transit, taxis)
    "tfl": "Transport", "uber": "Transport", "bolt": "Transport",
    "addison": "Transport", "national rail": "Transport",
    "trainline": "Transport", "heathrow": "Transport",
    "gatwick": "Transport", "hertz": "Transport", "avis": "Transport",
    "zipcar": "Transport", "lime": "Transport", # Lime e-scooter/bike
    # Travel & Accommodation (flights, hotels, airbnb)
    "airbnb": "Travel",
    "british airways": "Travel", "ryanair": "Travel", "easyjet": "Travel",
    "turkish airlines": "Travel", "jetblue": "Travel",
    "airlines": "Travel", "airways": "Travel",
    # Shopping
    "amazon": "Shopping", "asos": "Shopping", "zara": "Shopping",
    "h&m": "Shopping", "selfridges": "Shopping", "john lewis": "Shopping",
    "apple store": "Shopping", "nike": "Shopping", "adidas": "Shopping",
    "uniqlo": "Shopping", "batch baby": "Shopping",
    "weddingshop": "Shopping", "bold co": "Shopping",
    # Subscriptions & Bills
    "netflix": "Subscriptions", "spotify": "Subscriptions",
    "apple.com": "Subscriptions", "google one": "Subscriptions",
    "dropbox": "Subscriptions", "notion": "Subscriptions",
    "claude": "Subscriptions", "chatgpt": "Subscriptions",
    "openai": "Subscriptions", "linkedin": "Subscriptions",
    "islington council": "Subscriptions", "do energy": "Subscriptions",
    # Personal care
    "barber": "Personal care", "salon": "Personal care",
    "spa": "Personal care", "hairdresser": "Personal care",
    "true gents": "Personal care", "massage": "Personal care",
    # Rent / Housing
    "rent": "Rent",
    "psv and co": "Rent",    # property service charge
}

# Monzo category → our category
MONZO_CAT_MAP = {
    "eating out": "Dining", "restaurants": "Dining",
    "groceries": "Groceries", "supermarkets": "Groceries",
    "transport": "Transport", "travel": "Transport",
    "shopping": "Shopping", "clothing": "Shopping",
    "health": "Health", "medical": "Health",
    "entertainment": "Entertainment", "hobbies": "Entertainment",
    "bills": "Subscriptions", "subscriptions": "Subscriptions",
    "personal care": "Personal care",
    "housing": "Rent", "rent": "Rent",
    "general": "Misc", "savings": None,  # None = skip
}

# Skip these Monzo transaction types entirely
MONZO_SKIP_TYPES = {
    "pot transfer", "monzo-to-monzo", "bacs (direct credit)",
    "bank transfer",
}

# Skip Monzo transactions by merchant name (bill payments / internal transfers)
MONZO_SKIP_MERCHANTS = {
    "american express", "amex",          # Amex bill payment (already tracked in Amex CSVs)
    "monzo",                              # Internal monzo transfers
}

# Skip these Amex description patterns
AMEX_SKIP_PATTERNS = [
    r"^interest charge",
    r"^payment received",
    r"^balance transfer",
    r"^cash advance",
    r"^annual fee",
    r"^late payment fee",
    r"^returned payment",
]

# ─── CSV PARSERS ──────────────────────────────────────────────────────────────

def parse_monzo_csv(filepath):
    """
    Monzo export: Transaction ID, Date, Time, Type, Name, Emoji, Category,
                  Amount, Currency, Local amount, Local currency, Notes,
                  Address, Receipt, Description, Category split, Money Out, Money In
    Amount negative = spend; positive = income/refund (which we offset if card/flex/P2P).
    File 1 = current account; File 2 = Flex card.
    """
    rows = []
    filename = Path(filepath).name.lower()

    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            tx_type = row.get("Type", "").strip().lower()

            # Skip internal savings/pot transfers
            if tx_type in MONZO_SKIP_TYPES:
                continue
            if tx_type == "flex" and "early payment" in row.get("Notes and #tags", "").lower():
                continue  # Flex repayment, not a purchase

            # Skip bill payments to credit cards (not needed since Amex is gone, but keep for safety)
            merchant_raw = row.get("Name", "").strip().lower()
            if any(skip in merchant_raw for skip in MONZO_SKIP_MERCHANTS):
                continue

            raw_date = row.get("Date", "").strip()
            tx_date  = parse_date(raw_date)
            if not tx_date:
                continue

            raw_amt = row.get("Amount", "").strip().replace("£", "").replace(",", "")
            try:
                amount = float(raw_amt)
            except ValueError:
                continue

            # Determine if this is a spend (negative) or refund/offset (positive)
            is_spend = amount < 0
            is_offset = False

            if amount > 0:
                monzo_cat = row.get("Category", "").strip().lower()
                
                # Card/Flex payment refunds
                if tx_type in ("card payment", "flex", "refund"):
                    if monzo_cat not in ("savings", "transfers"):
                        is_offset = True
                
                # Peer-to-peer split payment paybacks (e.g. bills, dinners split)
                elif tx_type in ("monzo-to-monzo", "faster payment"):
                    # Only treat as offset if mapped to a valid category (and not Misc/Rent/Transfers)
                    mapped_cat = map_category(clean_merchant(row.get("Name", "").strip()), monzo_cat)
                    if mapped_cat and mapped_cat != "Misc" and mapped_cat in CATEGORIES and mapped_cat != "Rent":
                        is_offset = True

            if not is_spend and not is_offset:
                continue

            # Format: spending is stored as a positive number in Excel, refunds/offsets as a negative
            if is_offset:
                amount = -abs(round(amount, 2))
            else:
                amount = abs(round(amount, 2))

            merchant     = row.get("Name", "").strip() or row.get("Description", "").strip()
            merchant     = clean_merchant(merchant)
            monzo_cat    = row.get("Category", "").strip().lower()
            category     = map_category(merchant, monzo_cat)

            # Skip if category mapping says to skip (e.g. Savings)
            if category is None:
                continue

            tx_id = row.get("Transaction ID", "").strip()
            if not tx_id:
                tx_id = "monzo_" + tx_date.strftime("%Y%m%d") + "_" \
                        + re.sub(r"[^a-z0-9]", "", merchant.lower())[:10] \
                        + "_" + str(int(abs(amount) * 100))

            # Determine account label
            if "flex" in filename or tx_type == "flex":
                account = "Monzo Flex"
            else:
                account = "Monzo Current"

            rows.append({
                "date":     tx_date,
                "merchant": merchant,
                "amount":   amount,
                "category": category,
                "account":  account,
                "tx_id":    tx_id,
            })
    return rows


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def parse_date(s):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def clean_merchant(desc):
    """Strip Amex padding spaces and terminal location noise."""
    desc = re.sub(r"\s{2,}.*", "", desc)  # cut at double-space (Amex location suffix)
    return desc.strip().title()


def map_category(merchant, monzo_cat):
    m = merchant.lower()
    for keyword, cat in MERCHANT_MAP.items():
        if keyword in m:
            return cat
    if monzo_cat:
        return MONZO_CAT_MAP.get(monzo_cat, "Misc")
    return "Misc"


def fmt_gbp(n):
    return f"£{n:,.0f}"


def fmt_pct(n):
    return f"{n*100:.0f}%"


# ─── EXCEL UPDATER ────────────────────────────────────────────────────────────

TX_HEADERS = ["Date", "Merchant/Description", "Amount (£)", "Category",
              "Account", "Month", "Year", "Tx ID"]

# Colour palette (matching the tracker)
NAVY   = "0A1628"
DKBLUE = "0F2952"
MDBLUE = "1A4480"
GOLD   = "D4AF37"
WHITE  = "F8FAFC"
LGREY  = "94A3B8"
GREEN_FILL = "D1FAE5"
AMBER_FILL = "FEF3C7"
RED_FILL   = "FEE2E2"


def get_existing_ids(ws):
    ids = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[7]:
            ids.add(str(row[7]))
    return ids


def append_to_excel(transactions):
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found. Run build_tracker.py first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)

    if "Transactions" not in wb.sheetnames:
        print("ERROR: 'Transactions' sheet not found.")
        sys.exit(1)

    ws = wb["Transactions"]

    # Ensure header row is present
    if ws.max_row < 2 or not ws["A2"].value:
        for ci, h in enumerate(TX_HEADERS, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
            cell.fill      = PatternFill("solid", start_color=MDBLUE, end_color=MDBLUE)
            cell.alignment = Alignment(horizontal="center")

    existing_ids = get_existing_ids(ws)
    new_rows = [t for t in transactions if t["tx_id"] not in existing_ids]

    if not new_rows:
        print("  No new transactions to add.")
        return 0

    # Sort by date ascending
    new_rows.sort(key=lambda r: r["date"])

    next_row = max(ws.max_row + 1, 3)

    # Alternating row fills
    FILL_A = PatternFill("solid", start_color="0D2244", end_color="0D2244")
    FILL_B = PatternFill("solid", start_color=DKBLUE,  end_color=DKBLUE)

    for i, tx in enumerate(new_rows):
        r     = next_row + i
        fill  = FILL_A if i % 2 == 0 else FILL_B

        def sc(col, val, fmt=None, bold=False, color=WHITE, align="left"):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Arial", size=9, color=color, bold=bold)
            c.fill      = fill
            c.alignment = Alignment(horizontal=align)
            if fmt:
                c.number_format = fmt
            return c

        sc(1, tx["date"],     "dd/mm/yyyy", color="B0C4DE", align="center")
        sc(2, tx["merchant"], color=WHITE)
        sc(3, tx["amount"],   "£#,##0.00",  color="D4AF37", align="right")
        sc(4, tx["category"], color="93C5FD")
        sc(5, tx["account"],  color=LGREY)
        sc(6, tx["date"].month, color=LGREY, align="center")
        sc(7, tx["date"].year,  color=LGREY, align="center")
        sc(8, tx["tx_id"],    color="334155")

    wb.save(XLSX_PATH)
    print(f"  ✅ Added {len(new_rows)} new transactions.")
    return len(new_rows)


# ─── BUDGET ALERT ENGINE ──────────────────────────────────────────────────────

def get_budget_targets(wb):
    """Read budget targets from Config sheet."""
    cfg = wb["Config"]
    targets = {}
    # Config sheet: categories start at row 4, col B=category, col C=amount
    for row in cfg.iter_rows(min_row=4, max_row=14, values_only=True):
        if row[1] and row[2]:
            try:
                targets[str(row[1])] = float(row[2])
            except (ValueError, TypeError):
                continue
    return targets


def get_actuals(wb, month, year):
    """Sum actuals per category for a given month/year from Transactions."""
    ws = wb["Transactions"]
    actuals = {cat: 0.0 for cat in CATEGORIES}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        tx_date = row[0]
        if isinstance(tx_date, datetime):
            tx_date = tx_date.date()
        elif isinstance(tx_date, date):
            pass
        else:
            continue
        if tx_date.month == month and tx_date.year == year:
            cat = str(row[3]) if row[3] else "Misc"
            amt = float(row[2]) if row[2] else 0.0
            if cat in actuals:
                actuals[cat] += amt
    return actuals


def check_alerts(month=None, year=None, send_email=True):
    if not XLSX_PATH.exists():
        print("XLSX not found — skip alert check.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    today = date.today()
    month = month or today.month
    year  = year  or today.year

    targets = get_budget_targets(wb)
    actuals = get_actuals(wb, month, year)

    lifestyle_cats = [c for c in CATEGORIES if c != "Rent"]
    alerts = []
    for cat in lifestyle_cats:
        budget = targets.get(cat, 0)
        actual = actuals.get(cat, 0)
        if budget <= 0:
            continue
        pct = actual / budget
        if pct >= ALERT_THRESHOLD:
            alerts.append({
                "cat": cat, "actual": actual, "budget": budget,
                "pct": pct, "over": pct >= 1.0,
            })

    if not alerts:
        print("  ✅ All categories within budget.")
        return

    month_name = date(year, month, 1).strftime("%B %Y")
    print(f"\n  ⚠️  Budget alerts for {month_name}:")
    for a in alerts:
        icon = "🔴" if a["over"] else "⚠️"
        print(f"    {icon} {a['cat']}: {fmt_gbp(a['actual'])} / {fmt_gbp(a['budget'])} ({fmt_pct(a['pct'])})")

    if send_email and GMAIL_APP_PASSWORD:
        subject = ("🔴 Budget Alert: Overspend — " if any(a["over"] for a in alerts)
                   else "⚠️ Budget Alert: Approaching Limit — ") + month_name
        html = build_alert_html(alerts, month_name)
        send_gmail(subject, html)
    elif send_email:
        print("  (Email skipped — GMAIL_APP_PASSWORD not set)")


def build_month_brief(month=None, year=None, send_email=True):
    if not XLSX_PATH.exists():
        print("XLSX not found.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    today    = date.today()
    # Default: last month
    if not month:
        last = date(today.year, today.month, 1)
        from datetime import timedelta
        last = last - timedelta(days=1)
        month, year = last.month, last.year

    targets     = get_budget_targets(wb)
    actuals     = get_actuals(wb, month, year)
    prev_m      = month - 1 if month > 1 else 12
    prev_y      = year if month > 1 else year - 1
    prev_actuals = get_actuals(wb, prev_m, prev_y)

    lifestyle = [c for c in CATEGORIES if c != "Rent"]
    total_actual  = sum(actuals.get(c, 0) for c in lifestyle)
    total_budget  = sum(targets.get(c, 0) for c in lifestyle)
    total_prev    = sum(prev_actuals.get(c, 0) for c in lifestyle)

    flags = generate_flags(actuals, prev_actuals, targets, lifestyle)
    month_name = date(year, month, 1).strftime("%B %Y")

    html = build_brief_html(actuals, prev_actuals, targets, lifestyle,
                            total_actual, total_budget, total_prev,
                            flags, month_name)

    print(f"\n  📊 Monthly brief generated for {month_name}")
    print(f"     Total spend: {fmt_gbp(total_actual)} vs {fmt_gbp(total_budget)} budget")
    if flags:
        print(f"     Flags: {len(flags)}")

    if send_email and GMAIL_APP_PASSWORD:
        subject = f"💼 Budget Brief: {month_name} — CFO Analysis"
        send_gmail(subject, html)
    else:
        # Save to HTML file as fallback
        out = BASE_DIR / f"brief_{year}_{month:02d}.html"
        out.write_text(html, encoding="utf-8")
        print(f"     Brief saved to: {out.name}")


def generate_flags(actuals, prev_actuals, targets, cats):
    flags = []
    for cat in cats:
        actual = actuals.get(cat, 0)
        budget = targets.get(cat, 0)
        prev   = prev_actuals.get(cat, 0)
        if not budget:
            continue
        pct    = actual / budget
        mom    = (actual - prev) / prev if prev > 0 else None

        if pct > 1.0:
            flags.append({"type": "over", "cat": cat, "actual": actual,
                          "budget": budget, "pct": pct, "mom": mom,
                          "msg": f"{cat} over budget by {fmt_gbp(actual - budget)} (+{(pct-1)*100:.0f}%)",
                          "rec": f"Cut {cat.lower()} by {fmt_gbp(actual - budget)} next month."})
        elif mom and mom > 0.30:
            flags.append({"type": "spike", "cat": cat, "actual": actual,
                          "budget": budget, "pct": pct, "mom": mom,
                          "msg": f"{cat} up {mom*100:.0f}% MoM ({fmt_gbp(prev)} → {fmt_gbp(actual)})",
                          "rec": f"Verify if one-off. If recurring, raise {cat.lower()} budget by {fmt_gbp(actual - prev)}."})
        elif pct < 0.60 and actual > 0:
            flags.append({"type": "win", "cat": cat, "actual": actual,
                          "budget": budget, "pct": pct, "mom": mom,
                          "msg": f"{cat} at {pct*100:.0f}% — {fmt_gbp(budget - actual)} unspent",
                          "rec": f"Reallocate {fmt_gbp((budget - actual) // 2)} to savings or overspend category."})

    order = {"over": 0, "spike": 1, "win": 2}
    flags.sort(key=lambda f: order[f["type"]])
    return flags[:5]


# ─── EMAIL HTML ───────────────────────────────────────────────────────────────

def build_brief_html(actuals, prev_actuals, targets, cats,
                     total_actual, total_budget, total_prev,
                     flags, month_name):
    var     = total_actual - total_budget
    mom_pct = f"{(total_actual - total_prev) / total_prev * 100:+.1f}%" if total_prev > 0 else "N/A"
    var_col = "#22c55e" if var <= 0 else "#ef4444"

    rows_html = ""
    for cat in cats:
        actual = actuals.get(cat, 0)
        budget = targets.get(cat, 0)
        pct    = actual / budget if budget else 0
        bar    = min(int(pct * 100), 100)
        col    = "#ef4444" if pct > 1 else "#f59e0b" if pct > 0.8 else "#22c55e"
        status = "🔴 OVER" if pct > 1 else "⚠️ WARN" if pct > 0.8 else "✅ OK"
        rows_html += f"""
        <tr style="border-bottom:1px solid #1e3a5f">
          <td style="padding:8px 12px;color:#94a3b8;font-size:13px">{cat}</td>
          <td style="padding:8px 12px;text-align:right;color:#e2e8f0;font-family:monospace">{fmt_gbp(actual)}</td>
          <td style="padding:8px 12px;text-align:right;color:#64748b;font-family:monospace">{fmt_gbp(budget)}</td>
          <td style="padding:8px 12px;min-width:110px">
            <div style="background:#1e3a5f;border-radius:4px;height:8px">
              <div style="background:{col};border-radius:4px;height:8px;width:{bar}%"></div>
            </div>
          </td>
          <td style="padding:8px 12px;text-align:right;color:#94a3b8;font-size:12px">{pct*100:.0f}%</td>
          <td style="padding:8px 12px;text-align:center;font-size:12px">{status}</td>
        </tr>"""

    flags_html = ""
    for f in flags:
        icon = "🔴" if f["type"] == "over" else "⚠️" if f["type"] == "spike" else "💡"
        flags_html += f"""
        <tr style="border-bottom:1px solid #1e3a5f">
          <td style="padding:10px 12px;font-size:14px">{icon}</td>
          <td style="padding:10px 12px;color:#e2e8f0;font-size:13px">{f["msg"]}</td>
          <td style="padding:10px 12px;color:#94a3b8;font-size:12px">{f["rec"]}</td>
        </tr>"""
    if not flags_html:
        flags_html = '<tr><td colspan="3" style="padding:20px;color:#64748b;font-size:13px">No significant flags. Budget discipline maintained.</td></tr>'

    today_str = date.today().strftime("%A, %d %B %Y")
    return f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#0a1628;font-family:'Segoe UI',Arial,sans-serif">
<div style="max-width:700px;margin:0 auto;padding:24px">
  <div style="background:linear-gradient(135deg,#0f2952,#1a4480);border-radius:12px;padding:28px 32px;margin-bottom:20px">
    <div style="font-size:11px;color:#64748b;letter-spacing:2px;text-transform:uppercase;margin-bottom:6px">PRIVATE &amp; CONFIDENTIAL</div>
    <h1 style="margin:0 0 4px;font-size:24px;color:#f8fafc">Budget Brief</h1>
    <div style="font-size:15px;color:#94a3b8">{month_name} — CFO Analysis</div>
  </div>
  <div style="display:flex;gap:12px;margin-bottom:20px">
    <div style="flex:1;background:#0f2952;border:1px solid #1e3a5f;border-radius:10px;padding:18px 20px">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px">Total Spend</div>
      <div style="font-size:26px;font-weight:700;color:#f8fafc;font-family:monospace">{fmt_gbp(total_actual)}</div>
      <div style="font-size:12px;color:#94a3b8;margin-top:4px">vs {fmt_gbp(total_budget)} budget</div>
    </div>
    <div style="flex:1;background:#0f2952;border:1px solid #1e3a5f;border-radius:10px;padding:18px 20px">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px">Variance</div>
      <div style="font-size:26px;font-weight:700;color:{var_col};font-family:monospace">{"+" if var >= 0 else ""}{fmt_gbp(abs(var))}</div>
      <div style="font-size:12px;color:#94a3b8;margin-top:4px">{"Over" if var > 0 else "Under"} budget</div>
    </div>
    <div style="flex:1;background:#0f2952;border:1px solid #1e3a5f;border-radius:10px;padding:18px 20px">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:6px">vs Prior Month</div>
      <div style="font-size:26px;font-weight:700;color:#f8fafc;font-family:monospace">{mom_pct}</div>
      <div style="font-size:12px;color:#94a3b8;margin-top:4px">Month-on-month</div>
    </div>
  </div>
  <div style="background:#0f2952;border:1px solid #1e3a5f;border-radius:10px;margin-bottom:20px;overflow:hidden">
    <div style="padding:16px 20px;border-bottom:1px solid #1e3a5f">
      <h2 style="margin:0;font-size:14px;color:#f8fafc;text-transform:uppercase;letter-spacing:1px">Category Breakdown</h2>
    </div>
    <table style="width:100%;border-collapse:collapse">
      <thead><tr style="background:#0a1e3d">
        <th style="padding:8px 12px;text-align:left;font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:1px">Category</th>
        <th style="padding:8px 12px;text-align:right;font-size:11px;color:#64748b;text-transform:uppercase">Actual</th>
        <th style="padding:8px 12px;text-align:right;font-size:11px;color:#64748b;text-transform:uppercase">Budget</th>
        <th style="padding:8px 12px;font-size:11px;color:#64748b;text-transform:uppercase">Progress</th>
        <th style="padding:8px 12px;text-align:right;font-size:11px;color:#64748b;text-transform:uppercase">Used</th>
        <th style="padding:8px 12px;text-align:center;font-size:11px;color:#64748b;text-transform:uppercase">Status</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
  <div style="background:#0f2952;border:1px solid #1e3a5f;border-radius:10px;margin-bottom:20px;overflow:hidden">
    <div style="padding:16px 20px;border-bottom:1px solid #1e3a5f">
      <h2 style="margin:0;font-size:14px;color:#f8fafc;text-transform:uppercase;letter-spacing:1px">Analyst Flags &amp; Recommendations</h2>
    </div>
    <table style="width:100%;border-collapse:collapse">
      <thead><tr style="background:#0a1e3d">
        <th style="padding:8px 12px;width:30px"></th>
        <th style="padding:8px 12px;text-align:left;font-size:11px;color:#64748b;text-transform:uppercase">Finding</th>
        <th style="padding:8px 12px;text-align:left;font-size:11px;color:#64748b;text-transform:uppercase">Recommendation</th>
      </tr></thead>
      <tbody>{flags_html}</tbody>
    </table>
  </div>
  <div style="text-align:center;padding:16px;color:#334155;font-size:11px">
    Generated by DCR Budget System · {today_str}<br>
    Open DCR_Budget_Tracker.xlsx to review transactions in detail.
  </div>
</div></body></html>"""


def build_alert_html(alerts, month_name):
    rows_html = ""
    for a in alerts:
        icon  = "🔴" if a["over"] else "⚠️"
        label = "OVER BUDGET" if a["over"] else "APPROACHING"
        col   = "#ef4444" if a["over"] else "#f59e0b"
        rows_html += f"""
        <tr style="border-bottom:1px solid #1e3a5f">
          <td style="padding:10px 14px;font-size:14px">{icon}</td>
          <td style="padding:10px 14px;color:#e2e8f0;font-weight:600">{a["cat"]}</td>
          <td style="padding:10px 14px;text-align:right;font-family:monospace;color:#e2e8f0">{fmt_gbp(a["actual"])}</td>
          <td style="padding:10px 14px;text-align:right;font-family:monospace;color:#64748b">{fmt_gbp(a["budget"])}</td>
          <td style="padding:10px 14px;text-align:right;font-family:monospace;color:{col};font-weight:700">{fmt_pct(a["pct"])}</td>
          <td style="padding:10px 14px"><span style="background:{col};color:#fff;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700">{label}</span></td>
        </tr>"""
    return f"""<!DOCTYPE html><html><body style="margin:0;background:#0a1628;font-family:'Segoe UI',Arial,sans-serif">
<div style="max-width:640px;margin:0 auto;padding:24px">
  <div style="background:linear-gradient(135deg,#0f2952,#1a4480);border-radius:12px;padding:24px 28px;margin-bottom:18px">
    <h1 style="margin:0 0 4px;font-size:22px;color:#f8fafc">Budget Alert</h1>
    <div style="font-size:14px;color:#94a3b8">{month_name}</div>
  </div>
  <div style="background:#0f2952;border:1px solid #1e3a5f;border-radius:10px;overflow:hidden">
    <table style="width:100%;border-collapse:collapse">
      <thead><tr style="background:#0a1e3d">
        <th style="padding:8px 14px;width:30px"></th>
        <th style="padding:8px 14px;text-align:left;font-size:11px;color:#64748b;text-transform:uppercase">Category</th>
        <th style="padding:8px 14px;text-align:right;font-size:11px;color:#64748b;text-transform:uppercase">Actual</th>
        <th style="padding:8px 14px;text-align:right;font-size:11px;color:#64748b;text-transform:uppercase">Budget</th>
        <th style="padding:8px 14px;text-align:right;font-size:11px;color:#64748b;text-transform:uppercase">Used</th>
        <th style="padding:8px 14px;text-align:left;font-size:11px;color:#64748b;text-transform:uppercase">Status</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
  <div style="text-align:center;padding:14px;color:#334155;font-size:11px">DCR Budget System · Open DCR_Budget_Tracker.xlsx for full detail</div>
</div></body></html>"""


# ─── GMAIL SENDER ─────────────────────────────────────────────────────────────

def send_gmail(subject, html_body):
    if not GMAIL_APP_PASSWORD:
        print("  (Email not sent — GMAIL_APP_PASSWORD not configured)")
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = GMAIL_USER
        msg["To"]      = ALERT_EMAIL
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            s.sendmail(GMAIL_USER, ALERT_EMAIL, msg.as_string())
        print(f"  📧 Email sent: {subject}")
    except Exception as e:
        print(f"  ❌ Email failed: {e}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def clear_transactions():
    """Wipe all transaction data rows (keep headers) for a clean re-import."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["Transactions"]
    max_row = ws.max_row
    if max_row >= 3:
        ws.delete_rows(3, max_row - 2)
        print(f"  Cleared {max_row - 2} existing rows.")
    wb.save(XLSX_PATH)


def main():
    parser = argparse.ArgumentParser(description="DCR Budget Sync")
    parser.add_argument("--brief",  action="store_true", help="Generate + send monthly brief")
    parser.add_argument("--check",  action="store_true", help="Check budget alerts only")
    parser.add_argument("--reset",  action="store_true", help="Clear all transactions and re-import from scratch")
    parser.add_argument("--month",  type=int,  help="Month override (1-12)")
    parser.add_argument("--year",   type=int,  help="Year override (e.g. 2026)")
    args = parser.parse_args()

    if args.check:
        print("Checking budget alerts…")
        check_alerts(args.month, args.year)
        return

    if args.reset:
        print("Resetting transaction data…")
        clear_transactions()

    # ── 1. Collect all transactions from CSVs ──────────────────────────────
    all_tx = []
    errors = []

    print("Reading Monzo CSVs…")
    for f in sorted(MONZO_DIR.glob("*.csv")):
        try:
            rows = parse_monzo_csv(f)
            print(f"  {f.name}: {len(rows)} transactions")
            all_tx.extend(rows)
        except Exception as e:
            errors.append(f"Monzo {f.name}: {e}")

    if errors:
        print("\nWarnings:")
        for e in errors:
            print(f"  ⚠️  {e}")

    print(f"\nTotal transactions found: {len(all_tx)}")

    # ── 2. Deduplicate across all sources ─────────────────────────────────
    seen = {}
    deduped = []
    for tx in all_tx:
        if tx["tx_id"] not in seen:
            seen[tx["tx_id"]] = True
            deduped.append(tx)

    print(f"After dedup: {len(deduped)} unique transactions")

    # ── 3. Update Excel ────────────────────────────────────────────────────
    print(f"\nUpdating {XLSX_PATH.name}…")
    added = append_to_excel(deduped)

    # ── 4. Budget alerts ───────────────────────────────────────────────────
    print("\nChecking budget alerts…")
    check_alerts(args.month, args.year)

    # ── 5. Monthly brief if requested ─────────────────────────────────────
    if args.brief:
        print("\nBuilding monthly brief…")
        build_month_brief(args.month, args.year)

    print("\n✅ Sync complete.")
    if added > 0:
        print(f"   Open DCR_Budget_Tracker.xlsx to review {added} new transactions.")


if __name__ == "__main__":
    main()
