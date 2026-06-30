#!/usr/bin/env python3
"""
server.py — DCR Budget Tracker Local Server
===========================================
Serves the light-themed premium PWA mobile web app (index.html)
and provides API endpoints to read and write directly to
DCR_Budget_Tracker.xlsx using openpyxl.

Endpoints:
  GET  /                 → Serves index.html
  GET  /api/data         → Reads config & transactions from Excel, returns JSON
  POST /api/upload       → Handles CSV and PDF statements, parses and deduplicates
  POST /api/config       → Updates Start Date, End Date, and Starting Capital

Dependencies: openpyxl, pypdf (both standard on user's system)
"""

import os
import sys
import json
import re
import cgi
import traceback
from pathlib import Path
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, HTTPServer

import openpyxl
from openpyxl.styles import Font

# Re-use existing categories, merchant mappings, and helper logic from sync_budget.py
try:
    import sync_budget
    XLSX_PATH = sync_budget.XLSX_PATH
    CATEGORIES = sync_budget.CATEGORIES
    MERCHANT_MAP = sync_budget.MERCHANT_MAP
    MONZO_CAT_MAP = sync_budget.MONZO_CAT_MAP
    MONZO_SKIP_TYPES = sync_budget.MONZO_SKIP_TYPES
    MONZO_SKIP_MERCHANTS = sync_budget.MONZO_SKIP_MERCHANTS
    clean_merchant = sync_budget.clean_merchant
    map_category = sync_budget.map_category
    parse_date = sync_budget.parse_date
    append_to_excel = sync_budget.append_to_excel
    parse_monzo_csv = sync_budget.parse_monzo_csv
except ImportError:
    print("ERROR: sync_budget.py not found in directory. Ensure it is next to server.py.")
    sys.exit(1)

MONZO_DIR = Path(__file__).parent / "Monzo"
MONZO_DIR.mkdir(exist_ok=True)


def parse_monzo_pdf(filepath):
    """
    Parses a Monzo PDF statement using pypdf.
    Transaction line structure: DD/MM/YYYY Description Amount Balance
    E.g. 05/01/2026 True Gents Barber -29.70 1200.50
    """
    import pypdf
    rows = []
    
    # Regex to identify Monzo transaction lines:
    # starts with DD/MM/YYYY, followed by merchant description, followed by amount and running balance
    line_pattern = re.compile(r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(-?£?[\d,]+\.\d{2})\s+(£?[\d,]+\.\d{2})\s*$")

    try:
        reader = pypdf.PdfReader(filepath)
        for page_idx, page in enumerate(reader.pages):
            text = page.extract_text()
            if not text:
                continue
                
            for line in text.split("\n"):
                line = line.strip()
                match = line_pattern.match(line)
                if not match:
                    continue
                
                raw_date = match.group(1)
                merchant_raw = match.group(2).strip()
                raw_amt = match.group(3).strip().replace("£", "").replace(",", "")
                
                tx_date = parse_date(raw_date)
                if not tx_date:
                    continue
                
                try:
                    amount = float(raw_amt)
                except ValueError:
                    continue
                
                # Determine if transaction is a spend (negative) or refund/offset (positive)
                is_spend = amount < 0
                is_offset = False
                
                if amount > 0:
                    merchant_cleaned = clean_merchant(merchant_raw)
                    m_lower = merchant_cleaned.lower()
                    
                    # Card refunds typically have "refund" or "reversal"
                    if "refund" in m_lower or "reversal" in m_lower or "reversed" in m_lower:
                        is_offset = True
                    else:
                        # Or check if it maps to a valid spending category
                        mapped_cat = map_category(merchant_cleaned, "")
                        if mapped_cat and mapped_cat != "Misc" and mapped_cat in CATEGORIES and mapped_cat != "Rent":
                            is_offset = True
                
                if not is_spend and not is_offset:
                    continue
                
                # Format for Excel: Spends are positive, refunds are negative
                if is_offset:
                    amount = -abs(round(amount, 2))
                else:
                    amount = abs(round(amount, 2))
                
                merchant = clean_merchant(merchant_raw)
                category = map_category(merchant, "")
                
                if category is None:
                    continue
                
                # Create a reliable Tx ID for PDF lines
                tx_id = "monzo_" + tx_date.strftime("%Y%m%d") + "_" \
                        + re.sub(r"[^a-z0-9]", "", merchant.lower())[:10] \
                        + "_" + str(int(abs(amount) * 100))
                
                rows.append({
                    "date":     tx_date,
                    "merchant": merchant,
                    "amount":   amount,
                    "category": category,
                    "account":  "Monzo Current",
                    "tx_id":    tx_id,
                })
        print(f"  PDF Parser: successfully parsed {len(rows)} transactions from PDF.")
    except Exception as e:
        print(f"  PDF Parser Error: {e}")
        traceback.print_exc()
        
    return rows


def ensure_config_cells():
    """Initializes the passive observation config cells in Excel if empty."""
    if not XLSX_PATH.exists():
        return
        
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb["Config"]
    changed = False
    
    # Row 35: Start Date
    if not ws["C35"].value:
        ws["B35"] = "Observation Start Date"
        ws["C35"] = datetime.strptime("2026-07-01", "%Y-%m-%d").date()
        ws["C35"].number_format = 'yyyy-mm-dd'
        ws["B35"].font = Font(name="Arial", size=9, color="64748B")
        ws["C35"].font = Font(name="Arial", size=9, color="0000FF", bold=True)
        changed = True
        
    # Row 36: End Date
    if not ws["C36"].value:
        ws["B36"] = "Observation End Date"
        ws["C36"] = datetime.strptime("2026-08-31", "%Y-%m-%d").date()
        ws["C36"].number_format = 'yyyy-mm-dd'
        ws["B36"].font = Font(name="Arial", size=9, color="64748B")
        ws["C36"].font = Font(name="Arial", size=9, color="0000FF", bold=True)
        changed = True
        
    # Row 37: Starting Capital
    if not ws["C37"].value:
        ws["B37"] = "Starting Capital"
        ws["C37"] = 10000.0
        ws["C37"].number_format = '£#,##0.00'
        ws["B37"].font = Font(name="Arial", size=9, color="64748B")
        ws["C37"].font = Font(name="Arial", size=9, color="0000FF", bold=True)
        changed = True
        
    if changed:
        wb.save(XLSX_PATH)
        print("  Initialized Config worksheet rows 35-37.")


def read_excel_data():
    """Reads dates, starting capital, and full transactions list from Excel."""
    ensure_config_cells()
    
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    
    # 1. Load config values
    cfg = wb["Config"]
    start_date_val = cfg["C35"].value
    end_date_val = cfg["C36"].value
    starting_capital = cfg["C37"].value
    
    # Handle dates mapping safely
    if isinstance(start_date_val, (datetime, date)):
        start_date = start_date_val.strftime("%Y-%m-%d")
    else:
        start_date = str(start_date_val) if start_date_val else "2026-07-01"
        
    if isinstance(end_date_val, (datetime, date)):
        end_date = end_date_val.strftime("%Y-%m-%d")
    else:
        end_date = str(end_date_val) if end_date_val else "2026-08-31"
        
    try:
        starting_capital = float(starting_capital) if starting_capital is not None else 10000.0
    except (ValueError, TypeError):
        starting_capital = 10000.0
        
    # 2. Load transactions
    ws = wb["Transactions"]
    tx_list = []
    
    # Scan from Row 3 onwards
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 8 or not row[0]:
            continue
            
        t_date = row[0]
        if isinstance(t_date, (datetime, date)):
            date_str = t_date.strftime("%Y-%m-%d")
        else:
            # try parsing date string
            parsed = parse_date(str(t_date))
            if parsed:
                date_str = parsed.strftime("%Y-%m-%d")
            else:
                continue
                
        merchant = str(row[1]).strip() if row[1] else "Unknown"
        try:
            amount = float(row[2]) if row[2] is not None else 0.0
        except (ValueError, TypeError):
            amount = 0.0
            
        category = str(row[3]).strip() if row[3] else "Misc"
        account = str(row[4]).strip() if row[4] else "Monzo Current"
        tx_id = str(row[7]).strip() if row[7] else ""
        
        tx_list.append({
            "date":     date_str,
            "merchant": merchant,
            "amount":   amount,
            "category": category,
            "account":  account,
            "tx_id":    tx_id
        })
        
    # Sort transactions by date descending
    tx_list.sort(key=lambda t: t["date"], reverse=True)
    
    return {
        "start_date":       start_date,
        "end_date":         end_date,
        "starting_capital": starting_capital,
        "transactions":     tx_list
    }


def write_excel_config(start_date_str, end_date_str, starting_capital_float):
    """Writes the updated config parameters back to Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb["Config"]
    
    try:
        ws["C35"] = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    except Exception:
        pass
        
    try:
        ws["C36"] = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except Exception:
        pass
        
    try:
        ws["C37"] = float(starting_capital_float)
    except Exception:
        pass
        
    wb.save(XLSX_PATH)
    print("  Saved updated config values back to Excel.")


class BudgetRequestHandler(BaseHTTPRequestHandler):
    
    def log_message(self, format, *args):
        # Suppress noise logging for clinical aesthetics
        pass

    def send_cors_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, X-Requested-With")

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_cors_headers()
        self.end_headers()

    def do_GET(self):
        if self.path == "/" or self.path == "/index.html":
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_cors_headers()
            self.end_headers()
            
            # Read index.html
            html_path = Path(__file__).parent / "index.html"
            if html_path.exists():
                self.wfile.write(html_path.read_bytes())
            else:
                self.wfile.write(b"index.html not found. Please create the frontend file.")
                
        elif self.path == "/api/data":
            try:
                data = read_excel_data()
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps(data).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json")
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode("utf-8"))
                traceback.print_exc()
        else:
            # Fallback static server check
            file_path = Path(__file__).parent / self.path.lstrip("/")
            if file_path.exists() and file_path.is_file():
                self.send_response(200)
                if file_path.suffix == ".js":
                    self.send_header("Content-Type", "application/javascript")
                elif file_path.suffix == ".css":
                    self.send_header("Content-Type", "text/css")
                elif file_path.suffix == ".png":
                    self.send_header("Content-Type", "image/png")
                elif file_path.suffix == ".json":
                    self.send_header("Content-Type", "application/json")
                else:
                    self.send_header("Content-Type", "application/octet-stream")
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(file_path.read_bytes())
            else:
                self.send_response(404)
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(b"404 Not Found")

    def do_POST(self):
        if self.path == "/api/upload":
            try:
                # Process file upload
                ctype, pdict = cgi.parse_header(self.headers["Content-Type"])
                if ctype == "multipart/form-data":
                    # Parse form details
                    pdict["boundary"] = bytes(pdict["boundary"], "utf-8")
                    form = cgi.FieldStorage(
                        fp=self.rfile,
                        headers=self.headers,
                        environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers["Content-Type"]}
                    )
                    
                    if "file" not in form:
                        self.send_response(400)
                        self.send_header("Content-Type", "application/json")
                        self.send_cors_headers()
                        self.end_headers()
                        self.wfile.write(json.dumps({"error": "No file uploaded"}).encode("utf-8"))
                        return
                        
                    file_item = form["file"]
                    filename = file_item.filename
                    file_data = file_item.file.read()
                    
                    # Save to Monzo folder
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    safe_filename = f"uploaded_{timestamp}_{filename}"
                    dest_path = MONZO_DIR / safe_filename
                    dest_path.write_bytes(file_data)
                    print(f"  Upload: Statement saved raw to {dest_path}")
                    
                    # Parse depending on file extension
                    transactions = []
                    if filename.lower().endswith(".csv"):
                        transactions = parse_monzo_csv(dest_path)
                    elif filename.lower().endswith(".pdf"):
                        transactions = parse_monzo_pdf(dest_path)
                    else:
                        self.send_response(400)
                        self.send_header("Content-Type", "application/json")
                        self.send_cors_headers()
                        self.end_headers()
                        self.wfile.write(json.dumps({"error": "Unsupported file format (CSV or PDF only)"}).encode("utf-8"))
                        return
                    
                    # Append unique values to Excel
                    added = 0
                    if transactions:
                        added = append_to_excel(transactions)
                    
                    # Read updated database state
                    updated_data = read_excel_data()
                    updated_data["added_count"] = added
                    
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json")
                    self.send_cors_headers()
                    self.end_headers()
                    self.wfile.write(json.dumps(updated_data).encode("utf-8"))
                else:
                    self.send_response(400)
                    self.send_cors_headers()
                    self.end_headers()
                    self.wfile.write(b"Expected multipart/form-data")
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json")
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode("utf-8"))
                traceback.print_exc()
                
        elif self.path == "/api/config":
            try:
                content_length = int(self.headers["Content-Length"])
                post_data = self.rfile.read(content_length).decode("utf-8")
                params = json.loads(post_data)
                
                start_date = params.get("start_date")
                end_date = params.get("end_date")
                starting_capital = params.get("starting_capital")
                
                write_excel_config(start_date, end_date, starting_capital)
                
                # Fetch fresh values
                updated_data = read_excel_data()
                
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps(updated_data).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-Type", "application/json")
                self.send_cors_headers()
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode("utf-8"))
                traceback.print_exc()
        else:
            self.send_response(404)
            self.send_cors_headers()
            self.end_headers()


def run(port=8080):
    ensure_config_cells()
    server_address = ("", port)
    httpd = HTTPServer(server_address, BudgetRequestHandler)
    print(f"\n==================================================")
    print(f" DCR Spend Tracker & Observation Dev Server")
    print(f"==================================================")
    print(f" Serving web app on: http://localhost:{port}")
    print(f" Press Ctrl+C to stop...")
    print(f"==================================================\n")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
        sys.exit(0)


if __name__ == "__main__":
    run()
